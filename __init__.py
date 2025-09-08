import json, csv
from pathlib import Path
from openpyxl import Workbook, load_workbook
import yaml           # pip install pyyaml
from cryptography.fernet import Fernet

# ========== BASIC HELPERS ==========
def _load_json(path): 
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def _dump_json(data, path):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)

def _load_yaml(path):
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)

def _dump_yaml(data, path):
    with open(path, "w", encoding="utf-8") as f:
        yaml.safe_dump(data, f, sort_keys=False)

def _load_csv(path):
    with open(path, "r", encoding="utf-8") as f:
        return list(csv.DictReader(f))

def _dump_csv(data, path):
    if not data: return
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=data[0].keys())
        writer.writeheader(); writer.writerows(data)

def _load_xlsx(path):
    wb = load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows: return []
    headers, *body = rows
    return [dict(zip(headers, r)) for r in body]

def _dump_xlsx(data, path):
    wb = Workbook()
    ws = wb.active

    # if it's a dict, wrap it in a list
    if isinstance(data, dict):
        data = [data]

    if data and isinstance(data, list):
        ws.append(list(data[0].keys()))   # header row
        for row in data:
            ws.append(list(row.values()))
    wb.save(path)


# ========= CONVERSION CORE =========
_LOADERS = {
    ".json": _load_json,
    ".yml":  _load_yaml,
    ".yaml": _load_yaml,
    ".csv":  _load_csv,
    ".xlsx": _load_xlsx,
}

_DUMPERS = {
    ".json": _dump_json,
    ".yml":  _dump_yaml,
    ".yaml": _dump_yaml,
    ".csv":  _dump_csv,
    ".xlsx": _dump_xlsx,
}

def convert(input_path, output_path):
    """Convert any supported file (json/yaml/csv/xlsx) into another."""
    in_ext  = Path(input_path).suffix.lower()
    out_ext = Path(output_path).suffix.lower()
    if in_ext not in _LOADERS or out_ext not in _DUMPERS:
        raise ValueError(f"Unsupported formats: {in_ext}->{out_ext}")
    data = _LOADERS[in_ext](input_path)
    _DUMPERS[out_ext](data, output_path)

# ========= ENCRYPT / DECRYPT =========
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.backends import default_backend
import base64, os

def _derive_key(password: str, salt: bytes):
    kdf = PBKDF2HMAC(
        algorithm=hashes.SHA256(),
        length=32,
        salt=salt,
        iterations=390000,
        backend=default_backend(),
    )
    return base64.urlsafe_b64encode(kdf.derive(password.encode()))

def encrypt(path, password):
    salt = os.urandom(16)
    key = _derive_key(password, salt)
    f = Fernet(key)
    data = f.encrypt(open(path, "rb").read())
    # prepend salt so decrypt can reuse
    open(path + ".enc", "wb").write(salt + data)

def decrypt(path, password, out_path,out_name=None):
    raw = open(path, "rb").read()
    if len(raw) < 17:
        raise ValueError("Encrypted file is too short or corrupted.")
    salt, token = raw[:16], raw[16:]
    key = _derive_key(password, salt)
    f = Fernet(key)
    try:
        plain = f.decrypt(token)
    except Exception as e:
        raise ValueError(f"Decryption failed: {e}")
    # Use out_name if provided, else use out_path
    output_file = out_name if out_name else out_path
    open(output_file, "wb").write(plain)
